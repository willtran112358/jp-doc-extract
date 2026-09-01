"""Text extraction: PDF text-layer, CSV/XLSX, optional PaddleOCR and VLM."""

from __future__ import annotations

import csv
import re
from pathlib import Path

import pymupdf as fitz

from mapper import normalize_jp_text

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def extract_text_pdf_text_layer(pdf_path: Path) -> tuple[str, str]:
    doc = fitz.open(pdf_path)
    parts = [page.get_text("text") for page in doc]
    doc.close()
    text = normalize_jp_text("\n".join(parts))
    return (text, "text_layer") if len(text) >= 20 else (text, "ocr_needed")


def pdf_to_images(pdf_path: Path, max_pages: int = 5, dpi: int = 150) -> list[bytes]:
    doc = fitz.open(pdf_path)
    images: list[bytes] = []
    zoom = dpi / 72
    mat = fitz.Matrix(zoom, zoom)
    for i, page in enumerate(doc):
        if i >= max_pages:
            break
        pix = page.get_pixmap(matrix=mat, alpha=False)
        images.append(pix.tobytes("png"))
    doc.close()
    return images


def extract_text_paddle(pdf_path: Path, max_pages: int = 5) -> tuple[str, str]:
    try:
        from paddleocr import PaddleOCR  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError(
            "PaddleOCR not installed. Run: pip install -r requirements-optional.txt"
        ) from exc

    ocr = PaddleOCR(use_angle_cls=True, lang="japan", show_log=False)
    lines: list[str] = []
    for i, img_bytes in enumerate(pdf_to_images(pdf_path, max_pages=max_pages)):
        import tempfile

        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp.write(img_bytes)
            tmp_path = tmp.name
        result = ocr.ocr(tmp_path, cls=True)
        Path(tmp_path).unlink(missing_ok=True)
        if result and result[0]:
            for line in result[0]:
                if line and len(line) > 1:
                    lines.append(line[1][0])
        lines.append(f"[page:{i + 1}]")
    text = normalize_jp_text("\n".join(lines))
    return text, "paddle_ocr_jp"


def extract_text_vlm(pdf_path: Path, max_pages: int = 3) -> tuple[str, str]:
    import base64
    import os

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("Set ANTHROPIC_API_KEY for --mode vlm (see .env.example)")

    try:
        import anthropic  # type: ignore[import-untyped]
    except ImportError as exc:
        raise RuntimeError("pip install anthropic  (requirements-optional.txt)") from exc

    client = anthropic.Anthropic(api_key=api_key)
    images = pdf_to_images(pdf_path, max_pages=max_pages)
    content: list[dict] = [
        {
            "type": "text",
            "text": (
                "Extract all visible Japanese text from this document page. "
                "Preserve labels and numbers (dates, kWh, yen). Output plain text only."
            ),
        }
    ]
    for img in images:
        b64 = base64.standard_b64encode(img).decode("ascii")
        content.append(
            {
                "type": "image",
                "source": {"type": "base64", "media_type": "image/png", "data": b64},
            }
        )

    msg = client.messages.create(
        model=os.environ.get("ANTHROPIC_MODEL", "claude-sonnet-4-20250514"),
        max_tokens=4096,
        messages=[{"role": "user", "content": content}],
    )
    text = normalize_jp_text("".join(block.text for block in msg.content if hasattr(block, "text")))
    return text, "vlm_claude_vision"


def _read_csv_rows(path: Path) -> list[str]:
    raw = path.read_bytes()
    text = None
    for enc in ("utf-8-sig", "cp932", "shift_jis", "utf-8"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    text = normalize_jp_text(text)
    rows: list[str] = []
    for i, line in enumerate(csv.reader(text.splitlines())):
        if i > 120:
            break
        rows.append(" | ".join(line))
    return rows


def extract_text_csv(path: Path) -> tuple[str, str]:
    rows = _read_csv_rows(path)
    text = "\n".join(rows)
    header = rows[0] if rows else ""
    if re.search(r"使用量|電力量|kWh|排出|CO2|活動量|燃料", header, re.I):
        text = header + "\n" + text
    return text, "csv_parse"


def extract_text_xlsx(path: Path) -> tuple[str, str]:
    if load_workbook is None:
        return "", "xlsx_needs_openpyxl"
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[str] = []
    for ws in wb.worksheets[:3]:
        rows.append(f"[sheet:{ws.title}]")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 60:
                break
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append(" | ".join(cells))
    wb.close()
    return normalize_jp_text("\n".join(rows)), "xlsx_parse"


def extract_text(path: Path, mode: str = "auto") -> tuple[str, str]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        if mode == "paddle":
            return extract_text_paddle(path)
        if mode == "vlm":
            return extract_text_vlm(path)
        text, method = extract_text_pdf_text_layer(path)
        if method == "ocr_needed" and mode == "auto":
            return text, method
        return text, method
    if suffix == ".csv":
        return extract_text_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return extract_text_xlsx(path)
    raise ValueError(f"Unsupported type: {suffix}")
