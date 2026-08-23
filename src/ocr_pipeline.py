"""
CLAP AI Extraction POC — JP-optimized extract → Draft JSON
Supports: PDF (text layer), CSV (utf-8/cp932), XLSX
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

import pymupdf as fitz

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# After NFKC: numbers/ASCII normalized; keep JP keywords + flexible separators
NUM = r"([\d,]+(?:\.\d+)?)"
SEP = r"[\s　:：]*"
GAP = r"[\s　\n]{0,40}"

FIELD_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    (
        "company_name",
        re.compile(
            r"("
            r"(?:株式会社|有限会社|合同会社)[一-龥ぁ-んァ-ヶーA-Za-z0-9・･]{1,40}"
            r"|"
            r"[一-龥ぁ-んァ-ヶーA-Za-z0-9・･]{1,40}(?:株式会社|有限会社|合同会社)"
            r"|"
            r"[A-Za-z][A-Za-z0-9 .,&\-]{2,50}(?:Co\.,?\s*Ltd\.|Corporation|Inc\.)"
            r")"
        ),
    ),
    (
        "customer_name",
        re.compile(
            r"(?:お客様名|供給先名|納品先|検証対象|お客様)"
            + SEP
            + r"([^\n]{3,80})"
        ),
    ),
    (
        "invoice_no",
        re.compile(
            r"(?:請求書番号|伝票番号|納品書番号|計量票番号|文書番号|Invoice\s*No\.?)"
            + SEP
            + r"([A-Za-z0-9\-_/]+)",
            re.I,
        ),
    ),
    (
        "invoice_date",
        re.compile(
            r"(?:発行日|作成日|お支払期限)"
            + SEP
            + r"(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日?"
        ),
    ),
    (
        "invoice_date_iso",
        re.compile(r"(?:発行日|Date)" + SEP + r"(20\d{2})-(\d{1,2})-(\d{1,2})", re.I),
    ),
    (
        "billing_period",
        re.compile(
            r"(?:ご使用期間|対象期間|対象|請求)"
            + SEP
            + r"("
            r"20\d{2}年\d{1,2}月(?:分|度)?"
            r"|"
            r"20\d{2}年\d{1,2}月\d{1,2}日\s*[〜～\-~]\s*20\d{2}年\d{1,2}月\d{1,2}日"
            r"|"
            r"20\d{2}年\d{1,2}月\s*[〜～\-~]\s*20\d{2}年\d{1,2}月"
            r")"
        ),
    ),
    (
        "billing_period_loose",
        re.compile(r"(20\d{2}年\d{1,2}月(?:分|度)?)"),
    ),
    (
        "activity_amount",
        re.compile(
            r"(?:使用電力量|使用量|電力量|ガス使用量|数量|納入量|重量|Volume|Quantity)"
            + GAP
            + NUM
            + r"\s*(kWh|m3|m³|kg|t|トン|L|㎥)?",
            re.I,
        ),
    ),
    ("activity_kwh", re.compile(NUM + r"\s*kWh", re.I)),
    ("activity_m3", re.compile(NUM + r"\s*(?:m3|m³|㎥)", re.I)),
    (
        "amount_yen",
        re.compile(
            r"(?:ご請求額合計|ご請求金額|請求金額|税込合計|合計金額|請求金額サマリー|当月お買上額)"
            r"(?:[(（][^)）\n]{0,20}[)）])?"
            + GAP
            + r"[¥￥]?\s*"
            + NUM
            + r"\s*円?",
            re.I,
        ),
    ),
    (
        "emission_tco2e",
        re.compile(
            r"(?:Scope\s*[123]|総排出量|排出量|温室効果ガス)"
            + GAP
            + NUM
            + r"\s*(?:t-?CO2e?|トン)?",
            re.I,
        ),
    ),
    ("fiscal_year", re.compile(r"(FY\s*20\d{2}|FYE?\s*3/20\d{2}|20\d{2}年度|令和\d{1,2}年度)", re.I)),
]

DOC_TYPE_RULES: list[tuple[str, re.Pattern[str]]] = [
    ("electricity_invoice", re.compile(r"電気|電力|kWh|ご使用量のお知らせ|御請求書")),
    ("gas_invoice", re.compile(r"ガス|都市ガス|m³|m3|供給計量票")),
    ("shipping_invoice", re.compile(r"SHIPPING|海上輸送|Invoice", re.I)),
    ("coal_ticket", re.compile(r"石炭|煤炭|COAL", re.I)),
    ("milk_delivery", re.compile(r"生乳|納品書")),
    ("company_profile", re.compile(r"企業概要")),
    ("sustainability_report", re.compile(r"サステナビリティレポート")),
    ("verification_evidence", re.compile(r"検証|エビデンス|監査|指摘")),
]


def normalize_jp_text(text: str) -> str:
    """NFKC + JP whitespace/punctuation cleanup for stable regex."""
    if not text:
        return ""
    t = unicodedata.normalize("NFKC", text)
    t = t.replace("\u00a0", " ")
    t = t.replace("〜", "～")
    # collapse runs of spaces but keep newlines (field layout)
    t = re.sub(r"[^\S\n]+", " ", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()


def normalize_number(value: str) -> str:
    v = unicodedata.normalize("NFKC", value)
    return v.replace(",", "").replace("，", "").strip()


def classify_doc(text: str, name: str) -> str:
    n = name.lower()
    if any(k in name or k in n for k in ["電力", "electricity"]):
        return "electricity_invoice"
    if any(k in name or k in n for k in ["ガス", "gas"]):
        return "gas_invoice"
    if "coal" in n or "石炭" in name:
        return "coal_ticket"
    if "shipping" in n or "海上" in name:
        return "shipping_invoice"
    if any(k in name or k in n for k in ["生乳", "milk"]):
        return "milk_delivery"
    if any(k in name or k in n for k in ["profile", "企業概要"]):
        return "company_profile"
    if any(k in name or k in n for k in ["sustainability", "サステナ"]):
        return "sustainability_report"
    if any(k in name or k in n for k in ["evidence", "検証", "監査", "survey", "エビデンス"]):
        return "verification_evidence"
    if any(k in name or k in n for k in ["activity", "活動量", "ghg_calc", "排出", "emission", "fuel", "site_list"]):
        return "activity_table"
    blob = f"{name}\n{text[:1500]}"
    for label, pat in DOC_TYPE_RULES:
        if pat.search(blob):
            return label
    return "other"


def extract_text_pdf(pdf_path: Path) -> tuple[str, str]:
    doc = fitz.open(pdf_path)
    parts = [page.get_text("text") for page in doc]
    doc.close()
    text = normalize_jp_text("\n".join(parts))
    return (text, "text_layer") if len(text) >= 20 else (text, "ocr_needed")


def _read_csv_rows(path: Path) -> list[str]:
    raw = path.read_bytes()
    text = None
    for enc in ("utf-8-sig", "cp932", "shift_jis", "utf-8"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    text = normalize_jp_text(text)
    rows: list[str] = []
    for i, line in enumerate(csv.reader(text.splitlines())):
        if i > 120:
            break
        rows.append(" | ".join(line))
    return rows


def extract_text_csv(path: Path) -> tuple[str, str]:
    rows = _read_csv_rows(path)
    text = "\n".join(rows)
    header = rows[0] if rows else ""
    if re.search(r"使用量|電力量|kWh|排出|CO2|活動量|燃料", header, re.I):
        text = header + "\n" + text
    return text, "csv_parse"


def extract_text_xlsx(path: Path) -> tuple[str, str]:
    if load_workbook is None:
        return "", "xlsx_needs_openpyxl"
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[str] = []
    for ws in wb.worksheets[:3]:
        rows.append(f"[sheet:{ws.title}]")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 60:
                break
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append(" | ".join(cells))
    wb.close()
    return normalize_jp_text("\n".join(rows)), "xlsx_parse"


def extract_text(path: Path) -> tuple[str, str]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return extract_text_pdf(path)
    if suffix == ".csv":
        return extract_text_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return extract_text_xlsx(path)
    raise ValueError(f"Unsupported type: {suffix}")


def _set_field(fields: dict, evidence: list, name: str, value: str, snippet: str, conf: float) -> None:
    if name in fields:
        return
    fields[name] = {"value": value, "confidence": conf}
    evidence.append(
        {
            "field": name,
            "snippet": snippet[:140],
            "page": 1,
            "source": "regex_on_jp_normalized_text",
        }
    )


def map_fields(text: str) -> dict:
    text = normalize_jp_text(text)
    fields: dict = {}
    evidence: list = []

    for name, pat in FIELD_PATTERNS:
        m = pat.search(text)
        if not m:
            continue
        if name in {"invoice_date", "invoice_date_iso"} and m.lastindex and m.lastindex >= 3:
            value = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
            _set_field(fields, evidence, "invoice_date", value, m.group(0), 0.9)
        elif name == "activity_amount":
            value = normalize_number(m.group(1))
            unit = (m.group(2) or "").replace("トン", "t").replace("㎥", "m3")
            _set_field(fields, evidence, "activity_amount", value, m.group(0), 0.84)
            if unit:
                _set_field(fields, evidence, "activity_unit", unit, m.group(0), 0.82)
        elif name == "activity_kwh":
            _set_field(fields, evidence, "activity_amount", normalize_number(m.group(1)), m.group(0), 0.86)
            _set_field(fields, evidence, "activity_unit", "kWh", m.group(0), 0.88)
        elif name == "activity_m3":
            _set_field(fields, evidence, "activity_amount", normalize_number(m.group(1)), m.group(0), 0.86)
            _set_field(fields, evidence, "activity_unit", "m3", m.group(0), 0.88)
        elif name == "emission_tco2e":
            _set_field(fields, evidence, "emission_tco2e", normalize_number(m.group(1)), m.group(0), 0.72)
        elif name in {"billing_period", "billing_period_loose"}:
            value = next((g for g in m.groups() if g), m.group(0))
            _set_field(fields, evidence, "billing_period", value.strip(), m.group(0), 0.84)
        elif name == "amount_yen":
            _set_field(fields, evidence, "amount_yen", normalize_number(m.group(1)), m.group(0), 0.86)
        elif name == "customer_name":
            # stop at extra spaces / factory suffix noise lightly
            value = re.split(r"\s{2,}|　", m.group(1).strip())[0].strip()
            _set_field(fields, evidence, "customer_name", value, m.group(0), 0.88)
        else:
            value = m.group(1) if m.lastindex else m.group(0)
            if name in {"activity_amount", "amount_yen", "emission_tco2e"}:
                value = normalize_number(value)
            conf = 0.9 if name in {"company_name", "invoice_no"} else 0.8
            _set_field(fields, evidence, name, value.strip(), m.group(0), conf)

    return {"fields": fields, "evidence": evidence}


def build_draft(path: Path, text: str, method: str, mapped: dict) -> dict:
    doc_type = classify_doc(text, path.name)
    return {
        "poc": "clap-ai-ocr-poc",
        "track": "AI_Data_Extraction",
        "status": "draft",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "source_file": path.name,
        "source_dir": path.parent.name,
        "doc_type": doc_type,
        "extract_method": method,
        "jp_text_normalized": True,
        "warning": (
            None
            if method in {"text_layer", "csv_parse", "xlsx_parse"}
            else "Low text / missing parser — Phase-2 OCR or openpyxl required"
        ),
        "soft_warnings": [
            f"{k} confidence < 0.8"
            for k, v in mapped["fields"].items()
            if isinstance(v, dict) and v.get("confidence", 1) < 0.8
        ],
        "draft_ghg": mapped["fields"],
        "evidence_history": mapped["evidence"],
        "text_preview": text[:900],
        "next_step": "User review/edit → Confirm → SF Validation Rule → GHG Core (not in this POC)",
    }


def run(path: Path, out_dir: Path) -> Path:
    text, method = extract_text(path)
    mapped = map_fields(text)
    draft = build_draft(path, text, method, mapped)
    out_dir.mkdir(parents=True, exist_ok=True)
    safe = re.sub(r"[^\w.\-]+", "_", path.stem, flags=re.UNICODE)[:80]
    out_path = out_dir / f"{safe}_draft.json"
    out_path.write_text(json.dumps(draft, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


def iter_samples(root: Path) -> list[Path]:
    exts = {".pdf", ".csv", ".xlsx"}
    return sorted(p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in exts)


def run_batch(root: Path, out_dir: Path) -> Path:
    summary = []
    for path in iter_samples(root):
        try:
            out = run(path, out_dir)
            draft = json.loads(out.read_text(encoding="utf-8"))
            summary.append(
                {
                    "file": str(path.relative_to(root)),
                    "doc_type": draft["doc_type"],
                    "method": draft["extract_method"],
                    "fields": list(draft["draft_ghg"].keys()),
                    "field_count": len(draft["draft_ghg"]),
                    "evidence": len(draft["evidence_history"]),
                    "output": out.name,
                    "ok": True,
                }
            )
            print(
                f"OK  {path.parent.name}/{path.name} → "
                f"{draft['doc_type']} fields={len(draft['draft_ghg'])} evidence={len(draft['evidence_history'])}"
            )
        except Exception as exc:  # noqa: BLE001
            summary.append({"file": path.name, "ok": False, "error": str(exc)})
            print(f"ERR {path.name}: {exc}")

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "sample_root": str(root),
        "total": len(summary),
        "ok": sum(1 for s in summary if s.get("ok")),
        "jp_optimized": True,
        "results": summary,
    }
    out_dir.mkdir(parents=True, exist_ok=True)
    report_path = out_dir / "batch_summary.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nBatch: {report['ok']}/{report['total']} OK → {report_path}")
    return report_path


def main() -> None:
    parser = argparse.ArgumentParser(description="CLAP JP document → Draft JSON POC")
    parser.add_argument("input", nargs="?", default="samples/jp_electricity_invoice_sakura.pdf")
    parser.add_argument("-o", "--out", default="output")
    parser.add_argument("--batch", action="store_true", help="Process all PDF/CSV/XLSX under input dir")
    args = parser.parse_args()
    path = Path(args.input)
    if not path.exists():
        raise SystemExit(f"Not found: {path}")

    if args.batch or path.is_dir():
        run_batch(path if path.is_dir() else path.parent, Path(args.out))
        return

    out_path = run(path, Path(args.out))
    draft = json.loads(out_path.read_text(encoding="utf-8"))
    print("=== CLAP AI OCR POC (JP-optimized) ===")
    print(f"file     : {path}")
    print(f"doc_type : {draft['doc_type']}")
    print(f"method   : {draft['extract_method']}")
    print(f"status   : {draft['status']} (Draft / 未確定)")
    print(f"fields   : {list(draft['draft_ghg'].keys())}")
    for k, v in draft["draft_ghg"].items():
        print(f"  - {k}: {v.get('value')} (conf={v.get('confidence')})")
    print(f"evidence : {len(draft['evidence_history'])} snippets")
    print(f"output   : {out_path}")
    if draft.get("warning"):
        print(f"warning  : {draft['warning']}")


if __name__ == "__main__":
    main()
